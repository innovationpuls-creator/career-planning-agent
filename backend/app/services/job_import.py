from __future__ import annotations

import html
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import delete
from sqlalchemy.orm import Session

from app.models.job_posting import JobPosting


EXPECTED_HEADER = (
    "岗位名称",
    "地址",
    "薪资范围",
    "公司名称",
    "所属行业",
    "公司规模",
    "公司类型",
    "岗位详情",
    "更新日期",
    "公司详情",
)
DEFAULT_SOURCE_DIR = Path(r"C:\Users\yzh\Desktop\feature_map\行业数据")
BR_TAG_RE = re.compile(r"<br\s*/?>", re.IGNORECASE)
HTML_TAG_RE = re.compile(r"<[^>]+>")
LEADING_COMPANY_REPEAT_RE = re.compile(r"^(?P<name>.+?)(?P=name)+")


@dataclass(slots=True)
class ImportStats:
    file_count: int = 0
    total_rows: int = 0
    imported_rows: int = 0
    skipped_rows: int = 0


def normalize_cell(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def clean_address(value: object) -> str | None:
    text = normalize_cell(value)
    if text is None:
        return None
    return re.sub(r"-None$", "", text)


def clean_job_detail(value: object) -> str | None:
    text = normalize_cell(value)
    if text is None:
        return None

    text = BR_TAG_RE.sub("\n", text)
    text = HTML_TAG_RE.sub("", text)
    text = html.unescape(text)
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t\f\v]+", " ", text)
    text = re.sub(r"\n[ \t\f\v]+", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = text.strip()
    return text or None


def clean_company_detail(value: object, company_name: str | None) -> str | None:
    text = normalize_cell(value)
    if text is None:
        return None

    text = html.unescape(text)
    text = re.sub(r"\s+", " ", text).strip()
    if company_name:
        escaped_name = re.escape(company_name)
        text = re.sub(rf"^(?:{escaped_name})+", company_name, text)
    else:
        match = LEADING_COMPANY_REPEAT_RE.match(text)
        if match:
            text = f"{match.group('name')}{text[match.end():]}"
    return text or None


def row_to_payload(row: tuple[object, ...]) -> dict[str, str | None] | None:
    if not any(value is not None and str(value).strip() for value in row):
        return None

    values = list(row) + [None] * (10 - len(row))
    job_title, address, salary_range, company_name, industry, company_size, company_type, job_detail, _, company_detail = values[:10]

    normalized_job_title = normalize_cell(job_title)
    normalized_company_name = normalize_cell(company_name)
    normalized_industry = normalize_cell(industry)
    if not normalized_job_title or not normalized_company_name or not normalized_industry:
        return None

    return {
        "industry": normalized_industry,
        "job_title": normalized_job_title,
        "address": clean_address(address),
        "salary_range": normalize_cell(salary_range),
        "company_name": normalized_company_name,
        "company_size": normalize_cell(company_size),
        "company_type": normalize_cell(company_type),
        "job_detail": clean_job_detail(job_detail),
        "company_detail": clean_company_detail(company_detail, normalized_company_name),
    }


def import_job_postings(db: Session, source_dir: Path = DEFAULT_SOURCE_DIR) -> ImportStats:
    if not source_dir.exists():
        raise FileNotFoundError(f"Source directory does not exist: {source_dir}")

    files = sorted(source_dir.rglob("*.xlsx"))
    stats = ImportStats(file_count=len(files))

    db.execute(delete(JobPosting))
    db.flush()
    db.expunge_all()

    for file_path in files:
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            rows = sheet.iter_rows(values_only=True)
            header_row = next(rows, None)
            header = tuple(header_row[:10]) if header_row else ()
            if header != EXPECTED_HEADER:
                raise ValueError(f"Unexpected header in {file_path}: {header!r}")

            batch: list[JobPosting] = []
            for row in rows:
                if row is None:
                    continue
                stats.total_rows += 1
                payload = row_to_payload(row)
                if payload is None:
                    stats.skipped_rows += 1
                    continue
                batch.append(JobPosting(**payload))
                stats.imported_rows += 1

            if batch:
                db.add_all(batch)
        finally:
            workbook.close()

    db.commit()
    return stats
